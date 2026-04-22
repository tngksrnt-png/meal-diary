import openpyxl
import sys

io = sys.stdout
io.reconfigure(encoding='utf-8')

path = r'C:\dev\test\dashboard\ReNA_HR_Dashboard_260306_하이원리싸이클링.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

print(f'=== Workbook: {path} ===')
print(f'Sheets ({len(wb.sheetnames)}): {wb.sheetnames}')
print()

for name in wb.sheetnames:
    ws = wb[name]
    print(f'--- Sheet: {name}  (dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}) ---')
    rows = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows, start=1):
        cells = [('' if v is None else str(v)) for v in row]
        # trim trailing empties
        while cells and cells[-1] == '':
            cells.pop()
        if not cells:
            print(f'{i:>4}: <empty>')
        else:
            print(f'{i:>4}: ' + ' | '.join(cells))
    print()
